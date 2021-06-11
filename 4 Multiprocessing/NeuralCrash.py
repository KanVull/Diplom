import pickle
import numpy as np
import multiprocessing
from multiprocessing import shared_memory
from PyQt5 import QtCore, QtGui
from PyQt5.QtWidgets import (
    QApplication,
    QComboBox,
    QFileDialog,
    QMainWindow, 
    QDialog, 
    QDialogButtonBox,
    QMessageBox,
    QProgressBar,
    QSpinBox,
    QTextEdit,
    QVBoxLayout, 
    QHBoxLayout,
    QFormLayout, 
    QWidget,
    QPushButton,
    QAction, 
    QScrollArea, 
    QLabel, 
    QCheckBox,
    QLineEdit,
    QPlainTextEdit,
)
import os
import sys
import time
from math import ceil
from matplotlib.backends.backend_qt5 import NavigationToolbar2QT
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg
from matplotlib.figure import Figure
import matplotlib.pyplot as plt
import matplotlib
import datetime
import openpyxl

class NeuralCrach():
    def __init__(self):
        self.model = None
        self._name = 'NoName'
        self._sigma = None
        self._test_value = None

    def setName(self, name_of_model):
        self._name = name_of_model    
 
    def run(self):
        app = QApplication(sys.argv)
        ex = _NeuralCrashWindow(self)
        ex.show()
        app.exec_()   

class CheckingWidget_UI(object):
    checking_ending = QtCore.pyqtSignal(NeuralCrach)

    def setupUI(self, widget, neuralconfig):
        self.neuralconfig = neuralconfig
        self.layout = QVBoxLayout(widget)

        status = QFormLayout()
        self.error = QPlainTextEdit()
        self.error.setReadOnly(True)
        self.label_load_model = QLabel()
        self.label_load_model.setText('Checking')
        self.label_load_testdata = QLabel()
        self.label_load_testdata.setText('Checking')
        self.label_get_tested_values = QLabel()
        self.label_get_tested_values.setText('Checking')
        self.label_set_tested_values = QLabel()
        self.label_set_tested_values.setText('Checking')
        self.label_test_model = QLabel()
        self.label_test_model.setText('Checking')
        self.status_load_model = ''
        self.status_load_testdata = ''
        self.status_get_tested_values = ''
        self.status_set_tested_values = ''
        self.status_test_model = ''

        status.addRow('load_model',         self.label_load_model)
        status.addRow('load_testdata',      self.label_load_testdata)
        status.addRow('get_tested_values',  self.label_get_tested_values)
        status.addRow('set_tested_values',  self.label_set_tested_values)
        status.addRow('test_model',         self.label_test_model)

        self.layout.addLayout(status)
        self.layout.addWidget(self.error)
        self.layout.setContentsMargins(0,0,0,0)
        self.check_function_load_model()

    def errorprint(self, text, checked_functions):
        self.error.appendPlainText(checked_functions + '\n' + text + '\n\n\t--------\n\n')

    def update_status(self, label_name, status):
        if label_name == 'load_model':
            self.status_load_model = status
            self.label_load_model.setText(status)
        elif label_name == 'load_testdata':
            self.status_load_testdata = status
            self.label_load_testdata.setText(status)
        elif label_name == 'get_tested_values':
            self.status_get_tested_values = status
            self.label_get_tested_values.setText(status)
        elif label_name == 'set_tested_values':
            self.status_set_tested_values = status
            self.label_set_tested_values.setText(status)
        elif label_name == 'test_model':
            self.status_test_model = status
            self.label_test_model.setText(status)

    def update_label(self, label):
        if label.text().endswith('...'):
            label.setText('Checking.')
        else:
            label.setText(label.text() + '.')    

    def update_NeuralCrash(self, new_state):
        self.neuralconfig = new_state

    def check_function_load_model(self):
        self.threadAnim_load_model = AnimationTimer(0.3)
        self.threadWork_load_model = QtCore.QThread()
        self.QObject_load_model = Check_function(self.neuralconfig, 'load_model')
        self.QObject_load_model.moveToThread(self.threadWork_load_model)

        self.threadWork_load_model.started.connect(self.QObject_load_model.run)
        self.QObject_load_model.function_checked_signal.connect(self.update_status)
        self.QObject_load_model.error_signal.connect(self.errorprint)
        self.QObject_load_model.return_new_state.connect(self.update_NeuralCrash)
        self.QObject_load_model.done.connect(self.threadAnim_load_model.requestInterruption)
        self.QObject_load_model.done.connect(self.check_function_load_testdata)
        self.threadAnim_load_model.time_passes.connect(lambda: self.update_label(self.label_load_model))

        self.threadAnim_load_model.start()
        self.threadWork_load_model.start()

    def check_function_load_testdata(self):
        self.threadAnim_load_testdata = AnimationTimer(0.3)
        self.threadWork_load_testdata = QtCore.QThread()
        self.QObject_load_testdata = Check_function(self.neuralconfig, 'load_testdata')
        self.QObject_load_testdata.moveToThread(self.threadWork_load_testdata)

        self.threadWork_load_testdata.started.connect(self.QObject_load_testdata.run)
        self.QObject_load_testdata.function_checked_signal.connect(self.update_status)
        self.QObject_load_testdata.error_signal.connect(self.errorprint)
        self.QObject_load_testdata.return_new_state.connect(self.update_NeuralCrash)
        self.QObject_load_testdata.done.connect(self.threadAnim_load_testdata.requestInterruption)
        self.QObject_load_testdata.done.connect(self.check_function_get_tested_values)
        self.threadAnim_load_testdata.time_passes.connect(lambda: self.update_label(self.label_load_testdata))

        self.threadAnim_load_testdata.start()
        self.threadWork_load_testdata.start()

    def check_function_get_tested_values(self):
        self.threadAnim_get_tested_values = AnimationTimer(0.3)
        self.threadWork_get_tested_values = QtCore.QThread()
        self.QObject_get_tested_values = Check_function(self.neuralconfig, 'get_tested_values')
        self.QObject_get_tested_values.moveToThread(self.threadWork_get_tested_values)

        self.threadWork_get_tested_values.started.connect(self.QObject_get_tested_values.run)
        self.QObject_get_tested_values.function_checked_signal.connect(self.update_status)
        self.QObject_get_tested_values.error_signal.connect(self.errorprint)
        self.QObject_get_tested_values.return_new_state.connect(self.update_NeuralCrash)
        self.QObject_get_tested_values.done.connect(self.threadAnim_get_tested_values.requestInterruption)
        self.QObject_get_tested_values.done.connect(self.check_function_set_tested_values)
        self.threadAnim_get_tested_values.time_passes.connect(lambda: self.update_label(self.label_get_tested_values))

        self.threadAnim_get_tested_values.start()
        self.threadWork_get_tested_values.start()

    def check_function_set_tested_values(self):
        self.threadAnim_set_tested_values = AnimationTimer(0.3)
        self.threadWork_set_tested_values = QtCore.QThread()
        self.QObject_set_tested_values = Check_function(self.neuralconfig, 'set_tested_values')
        self.QObject_set_tested_values.moveToThread(self.threadWork_set_tested_values)

        self.threadWork_set_tested_values.started.connect(self.QObject_set_tested_values.run)
        self.QObject_set_tested_values.function_checked_signal.connect(self.update_status)
        self.QObject_set_tested_values.error_signal.connect(self.errorprint)
        self.QObject_set_tested_values.done.connect(self.threadAnim_set_tested_values.requestInterruption)
        self.QObject_set_tested_values.done.connect(self.check_function_test_model)
        self.threadAnim_set_tested_values.time_passes.connect(lambda: self.update_label(self.label_set_tested_values))

        self.threadAnim_set_tested_values.start()
        self.threadWork_set_tested_values.start()

    def check_function_test_model(self):
        self.threadAnim_test_model = AnimationTimer(0.3)
        self.threadWork_test_model = QtCore.QThread()
        self.QObject_test_model = Check_function(self.neuralconfig, 'test_model')
        self.QObject_test_model.moveToThread(self.threadWork_test_model)

        self.threadWork_test_model.started.connect(self.QObject_test_model.run)
        self.QObject_test_model.function_checked_signal.connect(self.update_status)
        self.QObject_test_model.error_signal.connect(self.errorprint)
        self.QObject_get_tested_values.return_new_state.connect(self.update_NeuralCrash)
        self.QObject_test_model.done.connect(self.threadAnim_test_model.requestInterruption)
        self.QObject_test_model.done.connect(self.end_checking)
        self.threadAnim_test_model.time_passes.connect(lambda: self.update_label(self.label_test_model))

        self.threadAnim_test_model.start()
        self.threadWork_test_model.start()

    def end_checking(self):
        status_tuple = (
            self.status_load_model,
            self.status_load_testdata,
            self.status_get_tested_values,
            self.status_set_tested_values,
            self.status_test_model,
        )
        for status in status_tuple:
            if status != 'OK':
                return     
        self.checking_ending.emit(self.neuralconfig)

class CheckingWidget(QWidget, CheckingWidget_UI):
    def __init__(self, neuralconfig, parent=None):
        super(CheckingWidget, self).__init__(parent) 
        self.setupUI(self, neuralconfig)


class TestConfigWidget_UI(object):
    done = QtCore.pyqtSignal(object)
    config = np.repeat([i/10000 for i in range(0,1001,20)], 1)
    config_changed = QtCore.pyqtSignal()

    def setupUI(self, widget):
        self.layout = QVBoxLayout(widget)
        
        self.config_layout = QFormLayout()

        self.num_processors = QSpinBox()
        self.num_processors.setMinimum(1)
        self.num_processors.setMaximum(self.getMaximumProcessorCount())
        self.num_processors.setValue(self.num_processors.maximum())
        label_processors = QLabel()
        label_processors.setText('Specify the number of processors used')
        self.config_layout.addRow(self.num_processors, label_processors)

        algorithms = GeneratingRandomAlgorithms()
        self.algorithmComboBox = QComboBox()
        self.algorithmComboBox.addItems(algorithms.names)
        algorithmLabel = QLabel()
        algorithmLabel.setText('Choose algorithm')
        self.config_layout.addRow(self.algorithmComboBox, algorithmLabel)

        self.layout.addLayout(self.config_layout)

        ## sigma widget
        self.sigmalayout = QVBoxLayout()
        
        ## 0 - start; 1 - stop; 2 - step; 3 - The number of tests at each step
        layouts_edits = [QVBoxLayout() for _ in range(4)]

        labels_edits = [QLabel() for _ in range(4)]
        labels_edits[0].setText('Start')
        labels_edits[1].setText('Stop')
        labels_edits[2].setText('Step')
        labels_edits[3].setText('NTES')
        labels_edits[3].setToolTip('The number of tests at each step')

        self.line_edits = [QLineEdit() for _ in range(4)]
        for i in range(3):
            self.line_edits[i].setToolTip('Int or float number\n. is delimiter')
        self.line_edits[3].setToolTip('Int number > 0')

        layout_sssc = QHBoxLayout()
        for i in range(4):
            layouts_edits[i].addWidget(labels_edits[i])
            layouts_edits[i].addWidget(self.line_edits[i])
            layout_sssc.addLayout(layouts_edits[i])
        
        self.line_edits[0].setText('0')
        self.line_edits[1].setText('0.005')
        self.line_edits[2].setText('0.0001')
        for i in range(4):
            self.line_edits[i].textEdited.connect(self.update_config)
            # self.line_edits[i].textEdited.connect(lambda: print('j'))
        self.line_edits[3].setText('1')    


        self.preview = QLabel()
        self.preview.setText('Don\'t work')

        self.sigmalayout.addLayout(layout_sssc)
        self.sigmalayout.addWidget(self.preview)

        self.config_changed.connect(self.setConfig)
        self.layout.addLayout(self.sigmalayout)
        self.layout.addStretch(0)

        ## sigma ended

        self.buttonstart = QPushButton()
        self.buttonstart.clicked.connect(self.onStarted)
        self.setConfig()
        self.buttonstart.setText('Start test')
        self.layout.addWidget(self.buttonstart)
        self.layout.setContentsMargins(0,0,0,0)  

    def getMaximumProcessorCount(self):
        return multiprocessing.cpu_count()    

    def setError(self, text):
        self.preview.setText('Config error | ' + text)
        self.config = None
        self.config_changed.emit()

    def update_config(self):
        numbers = [edit.text() for edit in self.line_edits]
        try:
            start = float(numbers[0])       
            stop = float(numbers[1])
            step = float(numbers[2])
            ntes = int(numbers[3])
        except ValueError:
            self.setError('Wrong numbers')
            return

        self.line_edits[3].setText(str(ntes))    
        if start > stop:
            self.setError('Start > Stop')
            return
        elif (stop - start) / step < 1:
            self.setError('Unavailable Step')
            return
        elif ntes < 1:
            self.setError('NTES must be over 0')
            return
        self.config = np.arange(start, stop, step)
        if ntes > 1:
            self.config = np.repeat(self.config, ntes)
        if len(self.config) > 8:
            self.preview.setText('[' + ', '.join(list(map(str, self.config[:3]))) + ', ..., ' + ', '.join(list(map(str, self.config[-3:]))) + ']')
        else:
            self.preview.setText('[' + ', '.join(list(map(str, self.config))) + ']') 
        self.config_changed.emit()       

    def setConfig(self):
        self.buttonstart.setEnabled(self.config is not None)

    def onStarted(self):
        data = {
            'algorithm': self.algorithmComboBox.currentText(),
            'config': self.config,
            'processors': int(self.num_processors.text())
        }
        self.done.emit(data) 

class TestConfigWidget(QWidget, TestConfigWidget_UI):
    def __init__(self, parent=None):
        super(TestConfigWidget, self).__init__(parent) 
        self.setupUI(self)


class TestingWidget_UI(object):
    done = QtCore.pyqtSignal(object)
    def setupUI(self, widget, data, neuralconfig):
        self.data = data
        self.neuralconfig = neuralconfig
        self.previous_pos = 0
        self.seconds_per_operation = 0
        self.seconds_per_operation_list = list()
        
        self.layout = QVBoxLayout(widget)
        self.label_average = QLabel()
        self.label_average.setText('time left')
        self.label_average.setAlignment(QtCore.Qt.AlignRight | QtCore.Qt.AlignBottom)
        self.layout.addWidget(self.label_average)

        self.progress_bar = QProgressBar()
        self.progress_bar.setMinimum(0)
        self.progress_bar.setMaximum(len(data['config']))
        self.layout.addWidget(self.progress_bar)
        self.layout.setContentsMargins(0,0,0,0)


        self.start_test()

    def start_test(self):
        self.threadTimer = AnimationTimer(0.5)
        self.shm = shared_memory.ShareableList([0], name='Progress')
        self.threadWork = Testing_function(self.data, self.neuralconfig, self.shm.shm.name)
        self.threadWork.done.connect(self.threadTimer.requestInterruption)
        self.threadWork.done.connect(self.load_graph)
        self.threadTimer.time_passes.connect(self.update_progress)

        self.threadWork.start()
        self.threadTimer.start()

    def getTime(self):
        if not self.seconds_per_operation_list:
            return 'time'
        if len(self.seconds_per_operation_list) > 100:
            self.seconds_per_operation_list.pop(0)
        average_time = np.average(self.seconds_per_operation_list)
        seconds_left = average_time * (self.progress_bar.maximum() - self.previous_pos) 
        m, s = divmod(int(seconds_left), 60)
        h, m = divmod(m, 60)
        if h != 0:
            return f'Average {h} hours'
        if m != 0:
            return f'Average {m} minutes'    
        return f'{s} seconds'

    def update_progress(self):
        current_pos = self.shm[0]
        tick = self.threadTimer.tick_time
        if current_pos != self.previous_pos:
            diff = current_pos - self.previous_pos
            if diff > 1:
                if self.seconds_per_operation != 0:
                    self.seconds_per_operation_list.append(self.seconds_per_operation)
                diff -= 1
                for _ in range(diff):
                    self.seconds_per_operation_list.append(tick / diff)
            else:
                self.seconds_per_operation_list.append(self.seconds_per_operation)
            self.seconds_per_operation = 0
            self.previous_pos = current_pos    
        else:
            self.seconds_per_operation += tick

        self.label_average.setText(self.getTime() + ' left')
        self.progress_bar.setValue(self.shm[0])

    def load_graph(self, data):
        del(self.shm)
        self.done.emit(data)

class TestingWidget(QWidget, TestingWidget_UI):
    def __init__(self, data, neuralconfig, parent=None):
        super(TestingWidget, self).__init__(parent) 
        self.setupUI(self, data, neuralconfig)


class MplCanvas(FigureCanvasQTAgg):
    def __init__(self, parent=None, width=5, height=4, dpi=300):
        self.fig = Figure(figsize=(width, height), dpi=dpi)
        self.axes = self.fig.add_subplot(111)
        super(MplCanvas, self).__init__(self.fig)

class ResultWidget_UI(object):
    another_test = QtCore.pyqtSignal()
    def setupUI(self, widget, data, test_value, name):
        self.data = data
        self.test_value = test_value
        self.test_time = datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        self.name = name

        self.layout = QVBoxLayout(widget)
        self.getXY_values()
        self.toolbar, self.canvas = self.getGraph()
        self.layout.addWidget(self.toolbar)
        self.layout.addWidget(self.canvas)

        button_save_excel = QPushButton()
        button_save_txt = QPushButton()
        button_another_test = QPushButton()

        button_save_excel.setText('Save to Excel')
        button_save_txt.setText('Save to txt')
        button_another_test.setText('Try another config')
        
        button_save_excel.clicked.connect(lambda: self.save('excel'))
        button_save_txt.clicked.connect(lambda: self.save('txt'))
        button_another_test.clicked.connect(lambda: self.another_test.emit())
        
        buttons_layout = QHBoxLayout()
        buttons_layout.addWidget(button_another_test)
        buttons_layout.addStretch(0)
        buttons_layout.addWidget(button_save_excel)
        buttons_layout.addWidget(button_save_txt)

        self.layout.addLayout(buttons_layout)
        self.layout.setContentsMargins(0,0,0,0)

    def getXY_values(self):
        XY_data = {}
        self.minX, self.minY, self.maxX, self.maxY = None, None, None, None
        for lst in self.data:
            for sigma_result in lst:
                if sigma_result[0] in XY_data:
                    XY_data[sigma_result[0]].append(sigma_result[1])
                else:
                    XY_data[sigma_result[0]] = [sigma_result[1]]
            del(sigma_result)        
        del(lst)            
        for key in XY_data.keys():
            # min max X
            if self.maxX is None:
                self.maxX = key
            elif key > self.maxX:
                self.maxX = key
            if self.minX is None:
                self.minX = key    
            elif key < self.minX:
                self.minX = key

            # counting Y
            if len(XY_data[key]) > 1:
                XY_data[key] = np.average(XY_data[key])

            # min max Y
            if self.maxY is None:
                self.maxY = XY_data[key] 
            elif XY_data[key]  > self.maxY:
                self.maxY = XY_data[key]
            if self.minY is None:
                self.minY = XY_data[key]     
            elif XY_data[key]  < self.minY:
                self.minY = XY_data[key] 

        self.data = XY_data                   

    def getGraph(self):
        matplotlib.use('Qt5Agg')

        sc = MplCanvas(self, width=5, height=5, dpi=100)

        sc.axes.plot(list(self.data.keys()), list(self.data.values()))
        sc.axes.set_xlabel('Sigma')
        sc.axes.set_ylabel('Test results')
        sc.axes.set_title(self.name + ' | ' + self.test_time)
        toolbar = NavigationToolbar2QT(sc, self)
        sc.fig.tight_layout()
        return toolbar, sc

    def save(self, save_type):
        data = {
            'name': self.name,
            'data': self.data,
            'time': self.test_time,
            'test_value': self.test_value,
        }
        saver = Saver(data, save_type)
        saver.save()

class ResultWidget(QWidget, ResultWidget_UI):
    def __init__(self, data, test_value, name, parent=None):
        super(ResultWidget, self).__init__(parent) 
        self.setupUI(self, data, test_value, name)

class Saver():
    def __init__(self, data, save_type):
        self._data = data
        self._save_type = save_type

    def save(self):
        if self._save_type == 'excel':
            self._save_to_excel()
        elif self._save_type == 'txt':
            self._save_to_txt()

    def _save_to_excel(self):
        wb = openpyxl.Workbook()
        ws = wb.worksheets[0]
        ws.cell(row=1, column=1).value = 'Name of model: ' + self._data['name']
        ws.cell(row=2, column=1).value = 'Tested: ' + self._data['time']
        ws.cell(row=3, column=1).value = 'Your model tested value: ' + str(self._data['test_value'])
        ws.cell(row=5, column=1).value = 'Sigma'
        ws.cell(row=5, column=2).value = 'Tested value'
        for row, key in enumerate(list(self._data['data'].keys())):
            ws.cell(row=row + 6, column=1).value = key
            ws.cell(row=row + 6, column=2).value = self._data['data'][key][0]


        filepath = QFileDialog.getSaveFileName(caption='Save your test result to excel file',
                                               directory=f'./',
                                               filter='Excel file (*.xlsx)')
        wb.save(filepath[0])

    def _save_to_txt(self):
        filepath = QFileDialog.getSaveFileName(caption='Save your test result to txt file',
                                               directory=f'./',
                                               filter='Text file (*.txt)')
        with open(filepath[0], 'w') as f:
            f.write('Name of model: ' + self._data['name'] + '\n')
            f.write('Tested: ' + self._data['time'] + '\n')
            f.write('Your model tested value: ' + str(self._data['test_value']) + '\n\n')
            f.write('Sigma\tTested value\n')
            for key, value in self._data['data'].items():
                f.write(str(key) + '\t' + str(value[0]) + '\n')


class AnimationTimer(QtCore.QThread):
    time_passes = QtCore.pyqtSignal()
    def __init__(self, tick_time):
        QtCore.QThread.__init__(self)
        self.tick_time = tick_time

    def run(self):
        while not self.isInterruptionRequested():
            self.time_passes.emit()
            time.sleep(self.tick_time)

class Custom_Exception(Exception):
    pass

class PickleProblem(Custom_Exception):
    '''raised when model is not pickleable'''
    pass

class NumpyProblem(Custom_Exception):
    '''raised when get_tested_data returns not a list of numpy arrays'''
    pass

class Check_function(QtCore.QObject):
    return_new_state = QtCore.pyqtSignal(NeuralCrach)
    function_checked_signal = QtCore.pyqtSignal(str, str)
    error_signal = QtCore.pyqtSignal(str, str)
    done = QtCore.pyqtSignal()

    def __init__(self, neuralconfig, name):
        QtCore.QObject.__init__(self)
        self.neuralconfig = neuralconfig
        self.name = name

    def run(self):
        try:
            if self.name == 'load_model':
                self.neuralconfig.load_model()
                try:
                    pickle.dumps(self.neuralconfig.model)
                except TypeError:
                    raise PickleProblem

            elif self.name == 'load_testdata':
                self.neuralconfig.load_testdata()

            elif self.name == 'get_tested_values':
                self.neuralconfig._basedata = self.neuralconfig.get_tested_values()
                if type(self.neuralconfig._basedata) is not list:
                    raise NumpyProblem
                for array in self.neuralconfig._basedata:
                    if type(array) is not np.ndarray:
                        raise NumpyProblem  

            elif self.name == 'set_tested_values':
                self.neuralconfig.set_tested_values(self.neuralconfig._basedata)

            elif self.name == 'test_model':
                self.neuralconfig._test_value = self.neuralconfig.test_model(self.neuralconfig.model, self.neuralconfig.testdata)

            self.function_checked_signal.emit(self.name, 'OK')    
        except PickleProblem:
            self.function_checked_signal.emit(self.name, 'ERROR')
            self.error_signal.emit('Your model is not pickleable.\nTry to pickle your model before launch the program', self.name)
        except NumpyProblem:
            self.function_checked_signal.emit(self.name, 'ERROR')
            self.error_signal.emit('This method must returns a list of numpy arrays (numpy.ndarray)', self.name)
        except Exception as e:
            self.function_checked_signal.emit(self.name, 'ERROR')
            self.error_signal.emit(str(e), self.name)

        self.return_new_state.emit(self.neuralconfig)  
        self.done.emit()      

class Testing_function(QtCore.QThread):
    done = QtCore.pyqtSignal(object)

    def __init__(self, data, neuralconfig, shm_name):
        QtCore.QObject.__init__(self)
        self.data = data
        self.algorithm = GeneratingRandomAlgorithms()
        self.neuralconfig = neuralconfig
        self.shm_name = shm_name

    def run(self):
        multiprocessExecution = MultiprocessExecution(self.data, self.neuralconfig, self.shm_name)
        data = multiprocessExecution.multiprocessingCalc()
        self.done.emit(data) 

class MultiprocessExecution():
    def __init__(self, data, neuralconfig, shm_name):
        self.data = data
        self.algorithm = GeneratingRandomAlgorithms()
        self.neuralconfig = neuralconfig
        self.shm_name = shm_name

    def _create_new_model(self, sigma):
        new_weights = self.algorithm.generate(self.neuralconfig._basedata, sigma, self.data['algorithm'])
        return self.neuralconfig.set_tested_values(new_weights)

    def _compute_models(self, sigmas):
        data = list()
        shm = shared_memory.ShareableList(name=self.shm_name)
        for sigma in sigmas:
            model = self._create_new_model(sigma)
            data.append((sigma, self.neuralconfig.test_model(model, self.neuralconfig.testdata)))
            del(model)
            shm[0] += 1
        return data  

    def multiprocessingCalc(self):
        sigmas = self.data['config']
        cpu_count = self.data['processors']
        process_step = ceil(len(sigmas) / cpu_count)
        list_in_data = list()
        for i in range(ceil(len(sigmas) / process_step)):
            end = i * process_step + process_step if i + process_step <= len(sigmas) else len(sigmas)
            start = i * process_step   
            list_in_data.append(sigmas[start:end])

        with multiprocessing.get_context("spawn").Pool(cpu_count) as p:
            data = p.map(self._compute_models, list_in_data) 
        return data

class GeneratingRandomAlgorithms():
    names = [
        'normal',
        'not_normal'
    ]

    def generate(self, weight, sigma, algorithm):
        if algorithm == 'normal':
            return self._random_numpy_normal(weight, sigma)

    def _random_numpy_normal(self, weight, sigma):
        return [np.random.normal(array, sigma) for array in weight]  


class _NeuralCrashWindowUI(object):
    def clearLayout(self, layout):
        while layout.count():
            child = layout.takeAt(0)
            if child.widget():
                child.widget().deleteLater()    

    def setupUI(self, MainWindow, neuralconfig):
        self.neuralconfig = neuralconfig
        self.MainWindow = MainWindow
        self.MainWindow.setWindowTitle('Test your neural')
        self.MainWindow.resize(500, 500)

        self.centralWidget = QWidget(self.MainWindow)
        self.centralWidgetLayout = QVBoxLayout(self.centralWidget)
        # self.centralWidgetLayout.setContentsMargins(0,0,0,0)
        self.create_checking_widget()
        

    def setnewcentralwidget(self, name, widget):
        self.clearLayout(self.centralWidgetLayout)
        label = QLabel()
        label.setText(name)
        label.setFont(QtGui.QFont('Arial', 12))
        label.setFixedHeight(30)
        self.centralWidgetLayout.addWidget(label)
        self.centralWidgetLayout.addWidget(widget)
        self.MainWindow.setCentralWidget(self.centralWidget)
        # self.MainWindow.setContentsMargins(0, 0, 0, 0)

    def create_checking_widget(self):
        checkingWidget = CheckingWidget(self.neuralconfig)
        checkingWidget.checking_ending.connect(self.create_test_config_widget)
        self.setnewcentralwidget('Checking override functions', checkingWidget)

    def create_test_config_widget(self, neuralconfig):
        self.MainWindow.resize(500, 180)
        self.neuralconfig = neuralconfig
        testConfigWidget = TestConfigWidget()
        testConfigWidget.done.connect(self.start_testing)
        self.setnewcentralwidget('Set test configuration', testConfigWidget)

    def start_testing(self, data):
        self.MainWindow.resize(500, 100)
        testingWidget = TestingWidget(data, self.neuralconfig)
        testingWidget.done.connect(self.load_graph_widget)
        self.setnewcentralwidget('Testing', testingWidget)

    def load_graph_widget(self, data):
        self.MainWindow.resize(700, 700)
        resultWidget = ResultWidget(data, self.neuralconfig._test_value, self.neuralconfig._name)
        resultWidget.another_test.connect(lambda: self.create_test_config_widget(self.neuralconfig))
        self.setnewcentralwidget('Result', resultWidget)    

class _NeuralCrashWindow(QMainWindow, _NeuralCrashWindowUI):
    def __init__(self, neuralconfig, parent=None):
        super(_NeuralCrashWindow, self).__init__(parent) 
        self.setupUI(self, neuralconfig)
     